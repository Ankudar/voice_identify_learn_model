import os
import numpy as np 
import pandas as pd
import tensorflow as tf
import matplotlib.pyplot as plt
import pathlib
import librosa.display
from tqdm import tqdm
# from sklearn.metrics import confusion_matrix
from sklearn.model_selection import train_test_split
import librosa
import  warnings
import sounddevice as sd
from scipy.io.wavfile import write
# from sklearn.metrics import plot_confusion_matrix
import seaborn as sns
from sklearn.metrics import classification_report
warnings.filterwarnings("ignore")

from tensorflow.keras.layers import Dense, Conv1D
from tensorflow.keras.layers import LeakyReLU, BatchNormalization, Flatten, MaxPooling1D, Input
from sincnet_tensorflow import SincConv1D, LayerNorm
# from spela.spectrogram import Spectrogram 
# from spela.melspectrogram import Melspectrogram
from tensorflow.keras.optimizers import Adam
from tensorflow.keras.utils import plot_model

import visualkeras
from PIL import ImageFont

from tensorflow.keras.callbacks import EarlyStopping, Callback
from tensorflow.keras.utils import to_categorical
import umap
import pickle

from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill
from openpyxl.utils import get_column_letter

from keras.layers import Dropout

from itertools import product
import datetime

data_dir = './train/data_set/'
spkr_id_file = './data_lists/speaker_id.scp'

spkr_id = {}

with open(spkr_id_file, "r", encoding = 'utf-8') as file:
    for line in file:
        key, value = line.strip().split(":", 1)
        spkr_id[int(key)] = value

def get_wav_paths(speaker):
    speaker_path = data_dir + speaker
    all_paths = [item for item in os.listdir(speaker_path)]
    return all_paths

speakers = [name for name in os.listdir(data_dir) if os.path.isdir(os.path.join(data_dir, name))]
paths = {speaker: get_wav_paths(speaker) for speaker in speakers}

def generate_training_data(speaker_paths, speaker, label):
    wavs, labels = [], []
    for i in tqdm(speaker_paths):
        if str(i).endswith('.wav'):
            wav = load_wav(i, speaker)
            wavs.append(wav)
            labels.append(label)
    return wavs,labels

def load_wav(wav_path,speaker):
    wav_path = data_dir + speaker + '/' + wav_path
    wav_data, _ = librosa.load(wav_path,sr=16000)
    wav_data = wav_data[:112000]
    wav_data = wav_data.tolist()
    return wav_data

labels = {}
with open('./data_lists/speaker_id.scp', 'r', encoding = 'utf-8') as file:
    for line in file:
        key, value = line.strip().split(':')
        labels[key] = value

for speaker, label in zip(speakers, labels):
    speaker_key = speaker
    label_key = speaker
    speaker_paths_key = speaker
    wavs, labels_dict = generate_training_data(paths[speaker_paths_key], speaker, label)
    

    if not os.path.exists('./data_lists/data_wav/'):
        # Если директории не существует, создаем ее
        os.makedirs('./data_lists/data_wav/')

    wav_file_path = f'./data_lists/data_wav/{speaker_key}_wavs.pkl'
    label_file_path = f'./data_lists/data_wav/{label_key}_labels.pkl'

    # Проверяем, существуют ли файлы
    if not os.path.exists(wav_file_path) and not os.path.exists(label_file_path):
        with open(wav_file_path, 'wb') as f:
            pickle.dump(wavs, f)
        with open(label_file_path, 'wb') as f:
            pickle.dump(labels_dict, f)


def read_data(file_path):
    with open(file_path, 'r', encoding = 'utf-8') as f:
        lines = f.readlines()
    
    data, labels = [], []
    for line in lines:
        path, label = line.strip().split(":")
        wav_data, _ = librosa.load(path, sr=16000)  # Load the audio file using librosa
        if len(wav_data) < 112000:  # If the audio file is shorter than 112000 samples
            wav_data = np.pad(wav_data, (0, 112000 - len(wav_data)))  # Pad it with zeros
        else:
            wav_data = wav_data[:112000]  # If it's longer, trim it
        data.append(wav_data)
        labels.append(int(label))

    return np.array(data), np.array(labels)  # Convert the list of arrays and labels to a numpy array

train_data, train_labels = read_data('./data_lists/train.scp')
test_data, test_labels = read_data('./data_lists/test.scp')

train_data = np.expand_dims(train_data, axis=-1)
test_data = np.expand_dims(test_data, axis=-1)
num_classes = len(set(spkr_id))
train_labels = to_categorical(train_labels, num_classes = num_classes)
test_labels = to_categorical(test_labels, num_classes = num_classes)

param_grid = {
    'DROPOUT': [0, 0.1, 0.2, 0.3, 0.4, 0.5],
    'LEARNING_RATE': [0.1, 0.01, 0.001, 0.0001, 0.00001, 0.0005],
    'EPOCHS': [1000]
}

combinations = list(product(*param_grid.values()))

for combination in combinations:
    start_time = datetime.datetime.now()
    params = dict(zip(param_grid.keys(), combination))
    DROPOUT = params['DROPOUT']
    EPOCHS = params['EPOCHS']
    LEARNING_RATE = params['LEARNING_RATE']

    sinc_layer = SincConv1D(N_filt=64,
                            Filt_dim=129,
                            fs=16000,
                            stride=16,
                            padding="SAME")

    inputs = Input((train_data.shape[1], 1)) 

    x = sinc_layer(inputs)
    x = LayerNorm()(x)

    x = LeakyReLU(alpha=0.2)(x)
    x = Dropout(DROPOUT)(x)
    x = MaxPooling1D(pool_size=2)(x)

    x = Conv1D(64, 3, strides=1, padding='valid')(x)
    x = BatchNormalization(momentum=0.05)(x)
    x = LeakyReLU(alpha=0.2)(x)
    x = Dropout(DROPOUT)(x)
    x = MaxPooling1D(pool_size=2)(x)

    x = Conv1D(64, 3, strides=1, padding='valid')(x)
    x = BatchNormalization(momentum=0.05)(x)
    x = LeakyReLU(alpha=0.2)(x)
    x = Dropout(DROPOUT)(x)
    x = MaxPooling1D(pool_size=2)(x)

    x = Conv1D(128, 3, strides=1, padding='valid')(x)
    x = BatchNormalization(momentum=0.05)(x)
    x = LeakyReLU(alpha=0.2)(x)
    x = Dropout(DROPOUT)(x)
    x = MaxPooling1D(pool_size=2)(x)

    x = Conv1D(128, 3, strides=1, padding='valid')(x)
    x = BatchNormalization(momentum=0.05)(x)
    x = LeakyReLU(alpha=0.2)(x)
    x = Dropout(DROPOUT)(x)
    x = MaxPooling1D(pool_size=2)(x)

    x = Flatten()(x)

    x = Dense(256)(x)
    x = BatchNormalization(momentum=0.05, epsilon=1e-5)(x)
    x = LeakyReLU(alpha=0.2)(x)
    x = Dropout(DROPOUT)(x)

    x = Dense(256)(x)
    x = BatchNormalization(momentum=0.05, epsilon=1e-5)(x)
    x = LeakyReLU(alpha=0.2)(x)
    x = Dropout(DROPOUT)(x)

    prediction = Dense(num_classes, activation='softmax')(x)
    model = tf.keras.models.Model(inputs=inputs, outputs=prediction)

    early_stop = EarlyStopping(monitor='val_loss', patience=5, verbose=1, mode='min')

    model.compile(optimizer=Adam(learning_rate=LEARNING_RATE,
                                beta_1=0.9,
                                beta_2=0.999,
                                epsilon=1e-07,
                                amsgrad=False,
                                name='Adam'),
                  loss='categorical_crossentropy',
                  metrics=['accuracy'])

    history = model.fit(x=train_data,
                        y=train_labels,
                        epochs=EPOCHS,
                        validation_data=(test_data,test_labels),
                        callbacks=[early_stop])

    loss = history.history['loss'][-1]
    accuracy = history.history['accuracy'][-1]
    val_loss = history.history['val_loss'][-1]
    val_accuracy = history.history['val_accuracy'][-1]

    

    end_time = datetime.datetime.now()
    execution_time = end_time - start_time
    hours, remainder = divmod(execution_time.seconds, 3600)
    minutes, seconds = divmod(remainder, 60)

    params.update({
        'start_time': start_time,
        'end_time': end_time,
        'total_time': f'{hours}:{minutes}:{seconds}',
        'DROPOUT': DROPOUT,
        'LEARNING_RATE': LEARNING_RATE,
        'EPOCHS': EPOCHS,
        'loss': loss,
        'accuracy': accuracy,
        'val_loss': val_loss,
        'val_accuracy': val_accuracy
    })

    params['start_time'] = params['start_time'].strftime('%d.%m.%Y %H:%M:%S')
    params['end_time'] = params['end_time'].strftime('%d.%m.%Y %H:%M:%S')
    new_df = pd.DataFrame(params, index=[0])

    try:
        df = pd.read_excel('statistics.xlsx')
    except FileNotFoundError:
        df = pd.DataFrame()

    df = pd.concat([df, new_df], ignore_index=True)
    df.to_excel('statistics.xlsx', index=False)