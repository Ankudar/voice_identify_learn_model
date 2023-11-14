import os
import numpy as np 
import pandas as pd
import tensorflow as tf
import matplotlib.pyplot as plt
import pathlib
import librosa.display
from tqdm import tqdm
from sklearn.model_selection import train_test_split

from sklearn.ensemble import IsolationForest
from sklearn.inspection import DecisionBoundaryDisplay

import librosa
import  warnings
import sounddevice as sd
from scipy.io.wavfile import write
import seaborn as sns
from sklearn.ensemble import IsolationForest
from sklearn.metrics import classification_report
warnings.filterwarnings("ignore")

from tensorflow.keras.layers import Dense, Conv1D
from tensorflow.keras.layers import LeakyReLU, BatchNormalization, Flatten, MaxPooling1D, Input
from sincnet_tensorflow import SincConv1D, LayerNorm
from tensorflow.keras.optimizers import Adam, RMSprop
from tensorflow.keras.utils import plot_model

import visualkeras
from PIL import ImageFont

from tensorflow.keras.callbacks import EarlyStopping, Callback, ModelCheckpoint
from tensorflow.keras.utils import to_categorical
import umap
import pickle

from openpyxl import load_workbook
from openpyxl.styles import Color, PatternFill
from openpyxl.utils import get_column_letter

from keras.layers import Dropout
import logging
import plotly.graph_objects as go

tf.get_logger().setLevel(logging.ERROR)

data_dir = './data_set/'
spkr_id_file = './data_lists/speaker_id.scp'

DROPOUT = 0 # для первого слоя 0
DROPOUT_CONV = 0  # для слоев Conv1D 0
DROPOUT_DENSE = 0.1  # для слоев Dense 0,1
BATCH_SIZE = 0
EPOCHS = 1000
LEARNING_RATE = 0.001 #0.0005 или 0.001
EARLY_STOP = 20

# print(os.listdir(data_dir))

spkr_id = {}

with open(spkr_id_file, "r", encoding = 'utf-8') as file:
    for line in file:
        key, value = line.strip().split(":", 1)
        spkr_id[int(key)] = value

# print(spkr_id)

def get_wav_paths(speaker):
    speaker_path = data_dir + speaker
    all_paths = [item for item in os.listdir(speaker_path)]
    return all_paths

speakers = [name for name in os.listdir(data_dir) if os.path.isdir(os.path.join(data_dir, name))]
paths = {speaker: get_wav_paths(speaker) for speaker in speakers}

def generate_training_data(speaker_paths, speaker, label):
    wavs, labels = [], []
    for i in tqdm(speaker_paths):
        if str(i).endswith('.wav'):
            wav = load_wav(i, speaker)
            wavs.append(wav)
            labels.append(label)
    return wavs,labels

def load_wav(wav_path,speaker):
    wav_path = data_dir + speaker + '/' + wav_path
    wav_data, _ = librosa.load(wav_path,sr=16000)
    # wav_data = wav_data[:112000]
    wav_data = wav_data.tolist()
    return wav_data

labels = {}
with open('./data_lists/speaker_id.scp', 'r', encoding = 'utf-8') as file:
    for line in file:
        key, value = line.strip().split(':')
        labels[key] = value

for speaker, label in zip(speakers, labels):
    speaker_key = speaker
    label_key = speaker
    speaker_paths_key = speaker
    wavs, labels_dict = generate_training_data(paths[speaker_paths_key], speaker, label)
    

    if not os.path.exists('./data_lists/data_wav/'):
        # Если директории не существует, создаем ее
        os.makedirs('./data_lists/data_wav/')

    wav_file_path = f'./data_lists/data_wav/{speaker_key}_wavs.pkl'
    label_file_path = f'./data_lists/data_wav/{label_key}_labels.pkl'

    # Проверяем, существуют ли файлы
    if not os.path.exists(wav_file_path) and not os.path.exists(label_file_path):
        with open(wav_file_path, 'wb') as f:
            pickle.dump(wavs, f)
        with open(label_file_path, 'wb') as f:
            pickle.dump(labels_dict, f)


def read_data(file_path):
    with open(file_path, 'r', encoding = 'utf-8') as f:
        lines = f.readlines()
    
    data, labels = [], []
    for line in lines:
        path, label = line.strip().split(":")
        wav_data, _ = librosa.load(path, sr=16000)  # Load the audio file using librosa
        # if len(wav_data) < 112000:  # If the audio file is shorter than 112000 samples
        #     wav_data = np.pad(wav_data, (0, 112000 - len(wav_data)))  # Pad it with zeros
        # else:
        #     wav_data = wav_data[:112000]  # If it's longer, trim it
        data.append(wav_data)
        labels.append(int(label))

    return np.array(data), np.array(labels)  # Convert the list of arrays and labels to a numpy array

train_data, train_labels = read_data('./data_lists/train.scp')
test_data, test_labels = read_data('./data_lists/test.scp')

# Expanding the dimension of the data
train_data = np.expand_dims(train_data, axis=-1)
test_data = np.expand_dims(test_data, axis=-1)

num_classes = len(set(spkr_id))
train_labels = to_categorical(train_labels, num_classes = num_classes)
test_labels = to_categorical(test_labels, num_classes = num_classes)

# print(train_data.shape)

sinc_layer = SincConv1D(N_filt=64,
                        Filt_dim=129,
                        fs=16000,
                        stride=16,
                        padding="SAME")

inputs = Input((train_data.shape[1], 1)) 

x = sinc_layer(inputs)
x = LayerNorm()(x)

x = LeakyReLU(alpha=0.2)(x)
x = Dropout(DROPOUT_CONV)(x) #DROPOUT
x = MaxPooling1D(pool_size=2)(x)

x = Conv1D(32, 3, strides=1, padding='valid')(x)
x = BatchNormalization(momentum=0.05)(x)
x = LeakyReLU(alpha=0.2)(x)
x = Dropout(DROPOUT_CONV)(x)
x = MaxPooling1D(pool_size=2)(x)

x = Conv1D(64, 3, strides=1, padding='valid')(x)
x = BatchNormalization(momentum=0.05)(x)
x = LeakyReLU(alpha=0.2)(x)
x = Dropout(DROPOUT_CONV)(x)
x = MaxPooling1D(pool_size=2)(x)

x = Conv1D(128, 3, strides=1, padding='valid')(x)
x = BatchNormalization(momentum=0.05)(x)
x = LeakyReLU(alpha=0.2)(x)
x = Dropout(DROPOUT_CONV)(x)
x = MaxPooling1D(pool_size=2)(x)

x = Conv1D(256, 3, strides=1, padding='valid')(x)
x = BatchNormalization(momentum=0.05)(x)
x = LeakyReLU(alpha=0.2)(x)
x = Dropout(DROPOUT_CONV)(x)
x = MaxPooling1D(pool_size=2)(x)

x = Flatten()(x)

x = Dense(256)(x)
x = BatchNormalization(momentum=0.05, epsilon=1e-5)(x)
x = LeakyReLU(alpha=0.2)(x)
x = Dropout(DROPOUT_CONV)(x) #DROPOUT_DENSE

x = Dense(128)(x)
x = BatchNormalization(momentum=0.05, epsilon=1e-5)(x)
x = LeakyReLU(alpha=0.2)(x)
x = Dropout(DROPOUT_CONV)(x) #DROPOUT_DENSE

prediction = Dense(num_classes, activation='softmax')(x)
model = tf.keras.models.Model(inputs=inputs, outputs=prediction)

early_stop = EarlyStopping(monitor='val_loss', patience = EARLY_STOP, verbose=1, mode='min')

class CustomCallback(tf.keras.callbacks.Callback):
    def __init__(self, checkpoint_dir):
        super(CustomCallback, self).__init__()
        self.checkpoint_dir = checkpoint_dir
        self.best_val_loss = float('inf')  # инициализируем с бесконечностью

    def on_epoch_end(self, epoch, logs=None):
        current_val_loss = logs.get('val_loss')
        if current_val_loss < self.best_val_loss:
            self.best_val_loss = current_val_loss
            if os.path.exists(self.checkpoint_dir):
                with open(os.path.join(self.checkpoint_dir, 'summ.txt'), 'w') as f:
                    f.write(f"DROPOUT: {DROPOUT}\n")
                    f.write(f"DROPOUT_CONV: {DROPOUT_CONV}\n")
                    f.write(f"DROPOUT_DENSE: {DROPOUT_DENSE}\n")
                    f.write(f"LEARNING_RATE: {LEARNING_RATE}\n")
                    f.write(f"loss: {logs['loss']}\n")
                    f.write(f"accuracy: {logs['accuracy']}\n")
                    f.write(f"val_loss: {logs['val_loss']}\n")
                    f.write(f"val_accuracy: {logs['val_accuracy']}\n")

# Используйте свой пользовательский обратный вызов вместе с ModelCheckpoint
model_checkpoint = ModelCheckpoint("final-model", save_weights_only=False, monitor='val_loss', mode='min', save_best_only=True)
custom_callback = CustomCallback("final-model")

model.compile(optimizer=RMSprop(learning_rate = LEARNING_RATE),
                            loss='categorical_crossentropy',
                            metrics=['accuracy'])

history = model.fit(x=train_data,
                    y=train_labels,
                    batch_size=BATCH_SIZE,  # Добавьте размер пакета здесь
                    epochs=EPOCHS,
                    validation_data=(test_data, test_labels),
                    callbacks=[early_stop, model_checkpoint, custom_callback])

# Evaluate the model
model.evaluate(test_data, test_labels, batch_size=BATCH_SIZE)

plt.plot(history.history['accuracy'])
plt.plot(history.history['val_accuracy'])
plt.title('model accuracy')
plt.ylabel('accuracy')
plt.xlabel('epoch')
plt.legend(['train', 'test'], loc='upper left')
plt.show()

plt.plot(history.history['loss'])
plt.plot(history.history['val_loss'])
plt.title('model loss')
plt.ylabel('loss')
plt.xlabel('epoch')
plt.legend(['train', 'test'], loc='upper left')
plt.show()

load_model = tf.keras.models.load_model("final-model")
y_pred = load_model.predict(test_data, batch_size=BATCH_SIZE)
# print(len(y_pred))

y_pred_labels = [np.argmax(y_pred[i]) for i in range(len(y_pred))]

# Получение представлений перед последним слоем
model_without_last_layer = tf.keras.models.Model(inputs=model.inputs, outputs=model.layers[-2].output)
features = model_without_last_layer.predict(test_data)

# Применение UMAP
reducer = umap.UMAP()
embedding = reducer.fit_transform(features)

test_labels_multiclass = np.argmax(test_labels, axis=1)
print(classification_report(test_labels_multiclass, y_pred_labels))

# Загрузка файла и создание словаря
spkr_id_dict = {}
with open('./data_lists/speaker_id.scp', 'r', encoding = 'utf-8') as f:
    for line in f:
        id, name = line.strip().split(':')
        spkr_id_dict[int(id)] = name

# Замена ID класса на имена в метках классов
class_names = [spkr_id_dict[i] for i in np.argmax(test_labels, axis=1)]

# Визуализация
fig = go.Figure(data=go.Scatter(
    x=embedding[:, 0],
    y=embedding[:, 1],
    mode='markers',
    marker=dict(
        size=5,
        color=np.argmax(test_labels, axis=1), #set color equal to a variable
        colorscale='Spectral', # one of plotly colorscales
        showscale=True
    ),
    text=[f'Class: {name}' for name in class_names], # метки классов
    hoverinfo='text'
))

fig.update_layout(title='speaker dataset', autosize=False,
                  width=1900, height=1000, # Здесь установите размеры вашего экрана
                  margin=dict(l=50, r=50, b=100, t=100, pad=10))

fig.show()